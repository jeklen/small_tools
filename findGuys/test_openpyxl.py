from openpyxl import Workbook
wb = Workbook()

# ws = wb.active
ws1 = wb.create_sheet("Mysheet") # insert at the end
ws2 = wb.create_sheet("Mysheet2", 0) # insert at the first position

wb.save('test.xlsx')
